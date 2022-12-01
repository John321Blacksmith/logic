# */scraping_tools/csv_manager.py

import csv
from openpyxl import Workbook, load_workbook
from .great_parser import DataFetcher as Df


class FileManager(Df, Workbook):
	def __init__(self):
		super().__init__()

	def write_to_the_text_file(self, file_name, content):
		if file_name:
			try:
				with open(file_name, mode='a', encoding='utf-8') as f:
					f.write(content)
			except FileNotFounderror:
				print(f'The file \'{file_name}\' not found.')

	def write_to_the_json_file(self, file_name, content):
		pass


	def write_to_the_csv_file(self,
							  file_name,
							  source,
							  item,
							  site_dict: dict):

		fields = ['title', 'integer', 'link', 'image']

		# open or create a csv file and create a writer object

		with open(file_name, mode='a', encoding='utf-8') as f:
			writer_object = csv.DictWriter(f, fieldnames=fields)

			content = Df.fetch_content(source, item, site_dict)
			objects = Df.structure_data(item, site_dict, content)

			writer_object.writeheader()

			for i in range(0, len(objects)):
				writer_object.writerow(objects[i])
			else:
				print(f'The data has been successfully loaded to the file \'{file_name}\'.')


	def write_to_the_excel_file(self,
								csv_file_name,
								cols_names,
								excel_file=None,
								ws_title=None):

		if excel_file:
			wb = load_workbook(filename=excel_file)
		else:
			excel_file = csv_file_name[:-4] + '.xlsx'
			wb = Workbook()

		# activate a workbook 
		work_sheet = wb.active

		# define a worksheet name
		if ws_title:
			work_sheet.title = ws_title

		try:
			with open(csv_file_name, mode='r', encoding='utf-8') as f:
				csv_reader = csv.reader(f, delimiter=',')

				line_count = 0
				for row in csv_reader:
					if line_count == 0:
						work_sheet.append(cols_names)
						line_count += 1
					else:

						if row != ['title', 'integer', 'link', 'image']:
							work_sheet.append(row)
							line_count += 1
				else:
					wb.save(excel_file)

			print(f'Created file \'{excel_file}\'.')

		except FileNotFoundError:
			print(f'The file \'{csv_file_name}\' not found.')
